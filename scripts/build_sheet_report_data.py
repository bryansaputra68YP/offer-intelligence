#!/usr/bin/env python3

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "work" / "google_sheet_export" / "amazon_report.xlsx"
OUTPUT = ROOT / "outputs" / "offer_chatbot" / "sheet_report_data.js"


TIER_SHEETS = {"Tier 1", "Tier 2", "Tier 3", "Tier 4", "BLACK TIER"}


def is_empty(value):
    return value is None or str(value).strip() == ""


def compact_text(value):
    text = str(value or "").strip()
    return re.sub(r"\s+", " ", text)


def display_value(cell):
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, str):
        return compact_text(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        fmt = str(cell.number_format or "")
        if "%" in fmt:
            return f"{value * 100:.2f}".rstrip("0").rstrip(".") + "%"
        if "$" in fmt:
            return f"${value:,.2f}".rstrip("0").rstrip(".")
        if abs(value - round(value)) < 0.0000001:
            return str(int(round(value)))
        return f"{value:.6f}".rstrip("0").rstrip(".")
    return compact_text(value)


def numeric_value(cell):
    value = cell.value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    if isinstance(value, str):
        text = value.strip().replace("$", "").replace(",", "").replace("%", "")
        try:
            number = float(text)
        except ValueError:
            return None
        return number / 100 if "%" in value else number
    return None


def trim_matrix(rows):
    while rows and all(is_empty(value) for value in rows[-1]):
        rows.pop()
    max_col = 0
    for row in rows:
        for idx, value in enumerate(row, 1):
            if not is_empty(value):
                max_col = max(max_col, idx)
    return [row[:max_col] for row in rows], max_col


def find_header_row(rows):
    for index, row in enumerate(rows):
        normalized = [compact_text(value).lower() for value in row]
        has_merchant = any(value in {"merchant id", "merchant name", "brand", "brand name"} for value in normalized)
        has_metrics = any(value in {"clicks", "conversion", "order count", "revenue", "backend epc"} for value in normalized)
        if has_merchant and has_metrics:
            return index
    return None


def row_to_dict(headers, values):
    output = {}
    for index, header in enumerate(headers):
        key = compact_text(header) or f"Column {index + 1}"
        value = values[index] if index < len(values) else ""
        output[key] = value
    return output


def sheet_summary_from_intro(intro_rows):
    metric_labels = {"brand count", "total clicks", "order count", "revenue", "avg conversion", "objective", "target"}
    for row_index, row in enumerate(intro_rows):
        labels = []
        for idx, value in enumerate(row):
            label = compact_text(value)
            if label.lower() in metric_labels:
                labels.append((idx, label))
        if len(labels) < 2 or row_index + 1 >= len(intro_rows):
            continue

        value_row = intro_rows[row_index + 1]
        cards = []
        for idx, label in labels:
            value = compact_text(value_row[idx] if idx < len(value_row) else "")
            if not value and idx + 1 < len(value_row) and is_empty(row[idx + 1] if idx + 1 < len(row) else ""):
                value = compact_text(value_row[idx + 1])
            if value:
                cards.append({"label": "Objective" if label == "Target" else label, "value": value})
        if cards:
            return cards

    cards = []
    for row_index, row in enumerate(intro_rows):
        for idx, value in enumerate(row):
            label = compact_text(value)
            if label.lower() not in metric_labels:
                continue
            found = ""
            for candidate in row[idx + 1 :]:
                if not is_empty(candidate):
                    found = compact_text(candidate)
                    break
            if found:
                cards.append({"label": "Objective" if label == "Target" else label, "value": found})
    seen = set()
    unique_cards = []
    for card in cards:
        if card["label"] in seen:
            continue
        seen.add(card["label"])
        unique_cards.append(card)
    return unique_cards


def parse_sheet(ws):
    rows = []
    numeric_rows = []
    for row in ws.iter_rows():
        rows.append([display_value(cell) for cell in row])
        numeric_rows.append([numeric_value(cell) for cell in row])
    rows, max_col = trim_matrix(rows)
    numeric_rows = [row[:max_col] for row in numeric_rows[: len(rows)]]
    header_index = find_header_row(rows)
    title = compact_text(rows[0][0]) if rows and rows[0] else ws.title

    if header_index is None:
        return {
            "name": ws.title,
            "title": title or ws.title,
            "kind": "grid",
            "introRows": rows[:12],
            "headers": [],
            "rows": [],
            "grid": rows,
            "summaryCards": [],
        }

    headers = [compact_text(value) or f"Column {idx + 1}" for idx, value in enumerate(rows[header_index])]
    table_rows = []
    for values, nums in zip(rows[header_index + 1 :], numeric_rows[header_index + 1 :]):
        if all(is_empty(value) for value in values):
            continue
        table_rows.append(row_to_dict(headers, values))

    return {
        "name": ws.title,
        "title": title or ws.title,
        "kind": "tier" if ws.title in TIER_SHEETS else "table",
        "introRows": rows[:header_index],
        "headers": headers,
        "rows": table_rows,
        "grid": [],
        "summaryCards": sheet_summary_from_intro(rows[:header_index]),
    }


def main():
    wb = load_workbook(INPUT, data_only=True, read_only=True)
    sheets = [parse_sheet(ws) for ws in wb.worksheets]
    payload = {
        "source": "Google Sheet export",
        "sourceUrl": "https://docs.google.com/spreadsheets/d/1Q8Ee_bf2sw-pVqJ64zNWeLsRYC3d8C042AMAFNGxwKs/edit",
        "generatedAt": "",
        "sheets": sheets,
        "tierSheets": [sheet["name"] for sheet in sheets if sheet["name"] in TIER_SHEETS],
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(f"window.SHEET_REPORT_DATA={json.dumps(payload, ensure_ascii=False, separators=(',', ':'))};\n", encoding="utf-8")
    print(json.dumps({
        "sheets": len(sheets),
        "tierSheets": payload["tierSheets"],
        "output": str(OUTPUT),
    }))


if __name__ == "__main__":
    main()
